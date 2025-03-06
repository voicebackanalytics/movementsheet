import streamlit as st
import pandas as pd
from io import BytesIO
import re
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# Streamlit App
st.title("Movement Sheet Web App")

# File uploader for Excel files
uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    package_file_path = r"Package File Movement sheet.xlsx"

    # Read the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file)
    package_df = pd.read_excel(package_file_path)
    
    required_columns = {"BookingStatus", "ServiceType", "ServiceTime", "ETA", "DepartureFlightNumber",
                            "ArrivalFlightNumber", "TransitFlightNumber", "ETD", "Origin", "Destination",
                            "PackageName", "Nationality", "TravelClass", "Remarks"}
    # Check for missing columns
    missing_columns = required_columns - set(df.columns)
    if missing_columns:
        st.error(f"Missing columns in uploaded file: {', '.join(missing_columns)}")
        st.stop()

    st.write("### Uploaded Data:")
    st.dataframe(df)

    # Package Mapping
    package_mapping = dict(zip(package_df["Package_Name"], package_df["Revised_Package_Name"]))


    # Step 2: Filter BookingStatus
    df = df[df["BookingStatus"].isin(["Completed", "PaymentCompleted"])]

    # Step 3: Adjust "ServiceType" for RoundTrip
    df.loc[(df["ServiceType"] == "RoundTrip") & (df["ServiceTime"] == df["ETA"]), "ServiceType"] = "Arrival"
    df.loc[(df["ServiceType"] == "RoundTrip") & (df["ServiceTime"] != df["ETA"]), "ServiceType"] = "Departure"

    # Step 4: Create "Flight No." column
    def get_flight_no(row):
        if row["ServiceType"] == "Departure" and row["DepartureFlightNumber"] not in ["NA", "", None]:
            return row["DepartureFlightNumber"]
        elif row["ServiceType"] == "Arrival" and row["ArrivalFlightNumber"] not in ["NA", "", None]:
            return row["ArrivalFlightNumber"]
        elif row["ServiceType"] == "Transit":
            return f"{row['ArrivalFlightNumber']} / {row['TransitFlightNumber']}"
        return ""

    df["Flight No."] = df.apply(get_flight_no, axis=1)

    # Step 5: Create "ETA/ETD" column
    df["ETA/ETD"] = df["ETD"].fillna(df["ETA"])

    # Step 6: Clean "Origin" and "Destination" columns
    def clean_location(value):
        return "" if pd.isna(value) else re.split(r"[,/]", value)[0].strip()

    df["Origin"] = df["Origin"].apply(clean_location)
    df["Destination"] = df["Destination"].apply(clean_location)

    # Step 7: Create "Orig/Dest" column
    def get_orig_dest(row):
        if row["ServiceType"] == "Arrival":
            return row["Origin"]
        elif row["ServiceType"] in ["Departure", "RoundTrip"]:
            return row["Destination"]
        elif row["ServiceType"] == "Transit":
            return f"{row['Origin']} / {row['Destination']}"
        return ""

    df["Orig/Dest"] = df.apply(get_orig_dest, axis=1)

    # Step 8: Create "Terminal" column
    def get_terminal(flight_no):
        match = re.search(r"-(\d{4})", str(flight_no))
        return "T1" if match and match.group(1).startswith("5") else "T2"

    df["Terminal"] = df["Flight No."].apply(get_terminal)

    # Step 9: Map "Package" column
    df["Package"] = df["PackageName"].apply(lambda x: package_mapping.get(x, x))

    # Step 10: Add empty "Profile" and "GSO" columns
    df["Profile"] = ""
    df["GSO"] = ""

    # Step 11: Create "Serial Number" column
    df.insert(0, "Sr No.", range(1, len(df) + 1))

    # Step 12: Remove duplicate values in "Nationality" and "Class of Travel"
    def remove_duplicates(text):
        return "" if pd.isna(text) else " : ".join(dict.fromkeys(text.split(" : ")))

    df["Nationality"] = df["Nationality"].apply(remove_duplicates)
    df["TravelClass"] = df["TravelClass"].apply(remove_duplicates)

    # Step 13: Rename columns
    rename_columns = {
   "Serial Number": "Sr No.", "ServiceTime": "Service Time", "ItenaryNumber": "Itinerary No.",
    "ServiceType": "Service", "Package": "Package", "Terminal": "Terminal",
    "GuestName": "GUEST NAME", "TotalGuest": "Total Guest", "Flight No.": "Flight No.",
    "Orig/Dest": "Orig/Dest", "ETA/ETD": "ETA/ETD", "TravelClass": "Class of Travel",
    "PlacardCountryCode": "Country Code", "PlacardContactNo": "Placard Guest Contact No.",
    "PlacardName": "Placard Guest Name", "Nationality": "Nationality", "Age": "Age",
    "BillingContactNo": "Booker Contact No.", "BillingEmail": "Email Id", "Remarks": "Remarks"
    }
    df = df.rename(columns=rename_columns)

    # Step 14: Remove rows where "Remarks" is "Cancelled"
    df = df[df["Remarks"] != "Cancelled"]

    # Step 15: Reorder columns
    column_order = ["Sr No.", "Service Time", "Itinerary No.", "Service", "Package", "Terminal", "Profile",
                    "GUEST NAME", "Total Guest", "Flight No.", "Orig/Dest", "ETA/ETD", "Class of Travel",
                    "Country Code", "Placard Guest Contact No.", "Placard Guest Name", "Nationality", "Age",
                    "Booker Contact No.", "Email Id", "Remarks", "GSO"]
    #df = df[column_order]
    
    # Check which columns are missing
    missing_cols = [col for col in column_order if col not in df.columns]
    if missing_cols:
        st.write("⚠️ Missing columns:", missing_cols)

    # Reorder only existing columns
    df = df[[col for col in column_order if col in df.columns]]

    # Step 16: Save DataFrame to an Excel file in memory
    output = BytesIO()

    
    
    # Step 16: Save to Excel with Bold Formatting
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        workbook = writer.book
        sheet = writer.sheets["Sheet1"]
        bold_font = Font(bold=True)
    
        # Apply Bold only on Package Column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
            for cell in row:
                if str(cell.value).strip() in map(str.strip, package_mapping.values()):
                    cell.font = bold_font

    st.write("✅ Data processing completed successfully!")
    st.write("### Processed Data:")
    st.write(df)

    # Step 17: Download Button
    st.download_button(
        label="Download Excel",
        data=output,
        file_name="modified_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
